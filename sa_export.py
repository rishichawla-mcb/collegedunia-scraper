"""Study Abroad — exports (SA tables only; never touches domestic data).

MEMORY: mirrors the streaming rewrite in export.py. Rows are pulled off the
sqlite cursor one at a time and the workbook uses openpyxl's write_only mode,
so peak RSS is flat regardless of how many programs have been scraped. The
previous version called cur.fetchall() and built a full in-memory workbook,
which is what OOM-kills a 2 GB host once sa_programs gets large.
"""
from __future__ import annotations

BUILD = "2026-07-23a"

import io
import json
from typing import Iterator, List, Optional, Sequence, Tuple

import sa_db

TABLES = ("sa_programs", "sa_universities", "sa_countries", "sa_program_exams",
          "sa_scholarships", "sa_university_rankings", "sa_university_courses",
          "sa_university_costs", "sa_university_nearby")

XLSX_MAX_ROWS = 1_048_575
_HEAVY_COLS = ("raw_json",)


def _headers_and_keep(cur, include_raw: bool) -> Tuple[List[str], List[int]]:
    headers = [d[0] for d in cur.description]
    keep = [i for i, h in enumerate(headers) if include_raw or h not in _HEAVY_COLS]
    return headers, keep


def _cell(v):
    """Same value coercion the previous implementation used."""
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v)
    return v


def _fetch(table, db_path=sa_db.SA_DB_PATH):
    """Back-compat helper: materialises a whole table. The exporters below no
    longer use it."""
    with sa_db.connect(db_path) as conn:
        cur = conn.execute(f"SELECT * FROM {table}")
        cols = [d[0] for d in cur.description]
        return cols, cur.fetchall()


def to_xlsx(tables: Sequence[str] = TABLES, db_path: str = sa_db.SA_DB_PATH,
            include_raw: bool = True, out_path: Optional[str] = None,
            max_rows_per_sheet: int = XLSX_MAX_ROWS):
    """One sheet per SA table. Streams; peak memory is independent of size.
    out_path=None -> return bytes (unchanged behaviour); otherwise write and
    return the path."""
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font

    wb = Workbook(write_only=True)
    bold = Font(bold=True)
    truncated: List[Tuple[str, int]] = []

    for t in tables:
        with sa_db.connect(db_path) as conn:
            try:
                cur = conn.execute(f"SELECT * FROM {t}")
            except Exception:  # noqa: BLE001  (table absent)
                continue
            headers, keep = _headers_and_keep(cur, include_raw)
            ws = wb.create_sheet(t[:31])
            hdr = []
            for i in keep:
                c = WriteOnlyCell(ws, value=headers[i])
                c.font = bold
                hdr.append(c)
            ws.append(hdr)
            n = 0
            for row in cur:                      # streamed, never fetchall()
                if n >= max_rows_per_sheet:
                    truncated.append((t, n))
                    break
                ws.append([_cell(row[i]) for i in keep])
                n += 1

    if not wb.sheetnames:
        wb.create_sheet("empty")
    if truncated:
        import sys
        for t, n in truncated:
            print(f"[sa_export] WARNING: sheet '{t}' truncated at {n:,} rows "
                  f"(Excel limit). Use CSV for the full table.", file=sys.stderr)

    if out_path:
        wb.save(out_path)
        return out_path
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _iter_csv(table: str, db_path: str, include_raw: bool) -> Iterator[str]:
    import csv
    buf = io.StringIO()
    w = csv.writer(buf)

    def flush() -> str:
        s = buf.getvalue()
        buf.seek(0)
        buf.truncate(0)
        return s

    with sa_db.connect(db_path) as conn:
        cur = conn.execute(f"SELECT * FROM {table}")
        headers, keep = _headers_and_keep(cur, include_raw)
        w.writerow([headers[i] for i in keep])
        yield flush()
        for row in cur:
            w.writerow(["" if row[i] is None else row[i] for i in keep])
            if buf.tell() > 262144:
                yield flush()
        tail = flush()
        if tail:
            yield tail


def _via_tempfile(write_chunks, suffix: str) -> bytes:
    """Spool to disk then read back once — 1x the payload in RAM instead of 2x.
    See the note in export.py._via_tempfile."""
    import os, tempfile
    fd, tmp = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    try:
        with open(tmp, "w", encoding="utf-8", newline="") as fh:
            for chunk in write_chunks:
                fh.write(chunk)
        with open(tmp, "rb") as fh:
            return fh.read()
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass


def to_csv(table, db_path=sa_db.SA_DB_PATH, include_raw: bool = True,
           out_path: Optional[str] = None):
    if out_path:
        with open(out_path, "w", encoding="utf-8", newline="") as fh:
            for chunk in _iter_csv(table, db_path, include_raw):
                fh.write(chunk)
        return out_path
    return _via_tempfile(_iter_csv(table, db_path, include_raw), ".csv")


def _iter_json(table: str, db_path: str, include_raw: bool) -> Iterator[str]:
    with sa_db.connect(db_path) as conn:
        cur = conn.execute(f"SELECT * FROM {table}")
        headers, keep = _headers_and_keep(cur, include_raw)
        names = [headers[i] for i in keep]
        yield "[\n"
        first = True
        for row in cur:
            obj = {n: row[i] for n, i in zip(names, keep)}
            yield ("" if first else ",\n") + " " + json.dumps(obj, ensure_ascii=False)
            first = False
        yield "\n]\n"


def to_json(table, db_path=sa_db.SA_DB_PATH, include_raw: bool = True,
            out_path: Optional[str] = None):
    if out_path:
        with open(out_path, "w", encoding="utf-8") as fh:
            for chunk in _iter_json(table, db_path, include_raw):
                fh.write(chunk)
        return out_path
    return _via_tempfile(_iter_json(table, db_path, include_raw), ".json")
