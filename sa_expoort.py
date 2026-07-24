"""Study Abroad — exports (SA tables only; never touches domestic data)."""
from __future__ import annotations

BUILD = "2026-07-23a"

import io
import json

import sa_db

TABLES = ("sa_programs", "sa_universities", "sa_countries", "sa_program_exams")


def _fetch(table, db_path=sa_db.SA_DB_PATH):
    with sa_db.connect(db_path) as conn:
        cur = conn.execute(f"SELECT * FROM {table}")
        cols = [d[0] for d in cur.description]
        return cols, cur.fetchall()


def to_xlsx(tables=TABLES, db_path=sa_db.SA_DB_PATH) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    wb.remove(wb.active)
    for t in tables:
        try:
            cols, rows = _fetch(t, db_path)
        except Exception:  # noqa: BLE001
            continue
        ws = wb.create_sheet(t[:31])
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([("" if v is None else (json.dumps(v) if isinstance(v, (dict, list)) else v))
                       for v in r])
    if not wb.sheetnames:
        wb.create_sheet("empty")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(table, db_path=sa_db.SA_DB_PATH) -> bytes:
    import csv
    cols, rows = _fetch(table, db_path)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(cols)
    for r in rows:
        w.writerow(["" if v is None else v for v in r])
    return buf.getvalue().encode("utf-8")
