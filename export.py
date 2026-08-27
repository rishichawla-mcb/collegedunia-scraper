"""
Export the scraped data from SQLite to xlsx / csv / json.

Each table (courses, colleges, offerings) can be exported. The offerings export
is the "joined" view linking courses to the colleges that offer them.

MEMORY: every exporter streams. Rows are pulled straight off the sqlite cursor
and written out one at a time — nothing calls fetchall(), and the xlsx writer
uses openpyxl's write_only mode so no Cell objects are retained. Peak RSS is
therefore flat (~50 MB) regardless of table size. The old implementation held
the whole table AND a full in-memory workbook, which cost ~19 MB of RSS per
1,000 offering rows and OOM-killed a 2 GB host on the "all tables" export.

Pass out_path=... to write straight to a file and get the path back (the
caller can then stream it to the browser without holding it in RAM); omit it
to get bytes, as before.
"""

from __future__ import annotations

BUILD = "2026-07-23a"  # keep in sync across app/db/scraper/export (header checks this)

import io
import json
import os
import tempfile
from typing import Iterator, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import db

TABLES = ("courses", "colleges", "offerings", "college_courses", "colleges_directory",
          "course_enrichment")

# Excel's hard ceiling is 1,048,576 rows INCLUDING the header row.
XLSX_MAX_ROWS = 1_048_575

# Columns dropped when include_raw=False. raw_json is the full source object and
# is typically 60-90% of the export's weight.
_HEAVY_COLS = ("raw_json",)


def _headers_and_keep(cur, include_raw: bool) -> Tuple[List[str], List[int]]:
    headers = [d[0] for d in cur.description]
    keep = [i for i, h in enumerate(headers) if include_raw or h not in _HEAVY_COLS]
    return headers, keep


def _fetch(table: str, db_path: str = db.DB_PATH) -> Tuple[List[str], List[tuple]]:
    """Back-compat helper: materialises a whole table. Kept for callers that
    genuinely want it — the exporters below no longer use it."""
    with db.connect(db_path) as conn:
        cur = conn.execute(f"SELECT * FROM {table}")
        headers = [d[0] for d in cur.description]
        rows = [tuple(r) for r in cur.fetchall()]
    return headers, rows


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------
def to_xlsx(tables: Sequence[str] = TABLES, db_path: str = db.DB_PATH,
            include_raw: bool = True, out_path: Optional[str] = None,
            max_rows_per_sheet: int = XLSX_MAX_ROWS):
    """Write one sheet per table. Streams; peak memory is independent of size.

    include_raw=False drops the raw_json column (much smaller file).
    out_path=None  -> return the xlsx bytes (default, unchanged behaviour).
    out_path=<p>   -> write to p and return p.

    A sheet is capped at Excel's row ceiling (XLSX_MAX_ROWS); if that bites, a
    warning naming the sheet is written to stderr (visible in the host's logs).
    A table that does not exist on this DB is skipped rather than raising.
    """
    wb = Workbook(write_only=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    truncated: List[Tuple[str, int]] = []

    for table in tables:
        with db.connect(db_path) as conn:
            try:
                cur = conn.execute(f"SELECT * FROM {table}")
            except Exception:  # noqa: BLE001  (table absent on an old DB)
                continue
            headers, keep = _headers_and_keep(cur, include_raw)
            ws = wb.create_sheet(title=table[:31])
            if headers:
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = f"A1:{get_column_letter(len(keep))}1"
                for j, i in enumerate(keep, start=1):
                    ws.column_dimensions[get_column_letter(j)].width = \
                        min(max(len(headers[i]) + 4, 12), 60)
                hdr = []
                for i in keep:
                    c = WriteOnlyCell(ws, value=headers[i])
                    c.font = header_font
                    c.fill = header_fill
                    hdr.append(c)
                ws.append(hdr)
            n = 0
            for row in cur:                      # streamed, never fetchall()
                if n >= max_rows_per_sheet:
                    truncated.append((table, n))
                    break
                ws.append([row[i] for i in keep])
                n += 1

    if not wb.sheetnames:
        wb.create_sheet("empty")
    if truncated:
        import sys
        for t, n in truncated:
            print(f"[export] WARNING: sheet '{t}' truncated at {n:,} rows "
                  f"(Excel limit). Use CSV for the full table.", file=sys.stderr)

    if out_path:
        wb.save(out_path)
        return out_path
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_xlsx_file(tables: Sequence[str] = TABLES, db_path: str = db.DB_PATH,
                 include_raw: bool = True, out_dir: Optional[str] = None,
                 name: str = "collegedunia_export.xlsx") -> str:
    """Convenience: build the workbook straight to a temp file and return its path."""
    out_dir = out_dir or tempfile.gettempdir()
    os.makedirs(out_dir, exist_ok=True)
    return to_xlsx(tables, db_path, include_raw=include_raw,
                   out_path=os.path.join(out_dir, name))


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def _iter_csv(table: str, db_path: str, include_raw: bool) -> Iterator[str]:
    import csv
    buf = io.StringIO()
    w = csv.writer(buf)

    def flush() -> str:
        s = buf.getvalue()
        buf.seek(0)
        buf.truncate(0)
        return s

    with db.connect(db_path) as conn:
        cur = conn.execute(f"SELECT * FROM {table}")
        headers, keep = _headers_and_keep(cur, include_raw)
        w.writerow([headers[i] for i in keep])
        yield flush()
        for row in cur:
            w.writerow([row[i] for i in keep])
            if buf.tell() > 262144:              # flush every ~256 KB
                yield flush()
        tail = flush()
        if tail:
            yield tail


def _via_tempfile(write_chunks, suffix: str) -> bytes:
    """Spool chunks to disk, then read back once. Joining chunks in memory costs
    ~2x the payload (a giant str, then its encoded copy); this costs 1x. Matters
    at export scale: a 330 MB CSV peaked at 677 MB RSS the naive way."""
    fd, tmp = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    try:
        with open(tmp, "w", encoding="utf-8", newline="") as fh:
            for chunk in write_chunks:
                fh.write(chunk)
        with open(tmp, "rb") as fh:
            return fh.read()
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass


def to_csv(table: str, db_path: str = db.DB_PATH, include_raw: bool = True,
           out_path: Optional[str] = None):
    """Stream one table to CSV. Returns bytes, or the path when out_path is given."""
    if out_path:
        with open(out_path, "w", encoding="utf-8", newline="") as fh:
            for chunk in _iter_csv(table, db_path, include_raw):
                fh.write(chunk)
        return out_path
    return _via_tempfile(_iter_csv(table, db_path, include_raw), ".csv")


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------
def _iter_json(table: str, db_path: str, include_raw: bool) -> Iterator[str]:
    with db.connect(db_path) as conn:
        cur = conn.execute(f"SELECT * FROM {table}")
        headers, keep = _headers_and_keep(cur, include_raw)
        names = [headers[i] for i in keep]
        yield "[\n"
        first = True
        for row in cur:
            obj = {n: row[i] for n, i in zip(names, keep)}
            yield ("" if first else ",\n") + "  " + json.dumps(obj, ensure_ascii=False)
            first = False
        yield "\n]\n"


def to_json(table: str, db_path: str = db.DB_PATH, include_raw: bool = True,
            out_path: Optional[str] = None):
    """Stream one table to a JSON array. Returns bytes, or the path when given."""
    if out_path:
        with open(out_path, "w", encoding="utf-8") as fh:
            for chunk in _iter_json(table, db_path, include_raw):
                fh.write(chunk)
        return out_path
    return _via_tempfile(_iter_json(table, db_path, include_raw), ".json")


# ---------------------------------------------------------------------------
# Analytical workbooks
# ---------------------------------------------------------------------------
def to_analytics_xlsx(db_path: str = db.DB_PATH, out_path: Optional[str] = None):
    """A summary workbook: pivots/aggregations for quick analysis. These are
    GROUP BY results (tiny), so pandas is safe here."""
    import pandas as pd
    sheets = {}
    with db.connect(db_path) as conn:
        sheets["Courses by stream"] = pd.read_sql_query(
            "SELECT stream_name AS stream, COUNT(*) AS courses, "
            "SUM(colleges_count) AS total_college_slots "
            "FROM courses WHERE stream_name<>'' GROUP BY stream_name "
            "ORDER BY courses DESC", conn)
        sheets["Courses by type"] = pd.read_sql_query(
            "SELECT course_type, level, COUNT(*) AS courses FROM courses "
            "GROUP BY course_type, level ORDER BY courses DESC", conn)
        try:
            sheets["Colleges by city"] = pd.read_sql_query(
                "SELECT city, COUNT(DISTINCT college_id) AS colleges, "
                "COUNT(*) AS offerings FROM offerings WHERE city<>'' "
                "GROUP BY city ORDER BY colleges DESC", conn)
            sheets["Offerings by stream"] = pd.read_sql_query(
                "SELECT c.stream_name AS stream, COUNT(*) AS offerings, "
                "ROUND(AVG(NULLIF(o.fees_amount,0)),0) AS avg_first_year_fee "
                "FROM offerings o JOIN courses c ON o.course_id=c.course_id "
                "GROUP BY c.stream_name ORDER BY offerings DESC", conn)
        except Exception:
            pass
    target = out_path or io.BytesIO()
    with pd.ExcelWriter(target, engine="openpyxl") as xl:
        wrote = False
        for name, df in sheets.items():
            if df is not None and not df.empty:
                df.to_excel(xl, sheet_name=name[:31], index=False)
                wrote = True
        if not wrote:
            pd.DataFrame({"info": ["no data yet"]}).to_excel(xl, sheet_name="empty", index=False)
    return out_path if out_path else target.getvalue()


_MASTER_SQL = (
    "SELECT o.course_name, o.course_type, o.level, o.college_name, o.city, "
    "o.state_id, o.fees_amount AS first_year_fee_inr, o.fees_text, o.eligibility, "
    "o.exam_name, o.duration, o.ranking_rank, o.ranking_agency, o.course_rating, "
    "o.reviews_count, o.cutoff_exam, o.cutoff_value, o.admission_start, "
    "o.admission_end, c.website, c.email, c.phone, "
    "c.rating_value AS college_rating, c.rating_count AS college_reviews, "
    "c.address, o.university_link "
    "FROM offerings o LEFT JOIN colleges c ON o.college_id=c.college_id "
    "LIMIT ?"
)


def to_master_xlsx(db_path: str = db.DB_PATH, out_path: Optional[str] = None,
                   limit: int = 200000):
    """One analytical sheet: each offering joined with college enrichment.
    Streamed the same way as to_xlsx (the old pandas version materialised the
    whole 200k-row frame AND a full in-memory workbook)."""
    wb = Workbook(write_only=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    with db.connect(db_path) as conn:
        cur = conn.execute(_MASTER_SQL, (int(min(limit, XLSX_MAX_ROWS)),))
        headers = [d[0] for d in cur.description]
        first = cur.fetchone()
        ws = wb.create_sheet(title="master")
        if first is None:
            # Match the previous implementation's empty-state sheet exactly.
            ws.append([WriteOnlyCell(ws, value="info")])
            ws.append(["no offerings yet"])
        else:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
            for j, h in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(j)].width = min(max(len(h) + 4, 12), 60)
            hdr = []
            for h in headers:
                c = WriteOnlyCell(ws, value=h)
                c.font = header_font
                c.fill = header_fill
                hdr.append(c)
            ws.append(hdr)
            ws.append(list(first))
            for row in cur:
                ws.append(list(row))
    if out_path:
        wb.save(out_path)
        return out_path
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "collegedunia_export.xlsx"
    to_xlsx(out_path=out)
    print(f"Wrote {out}")
